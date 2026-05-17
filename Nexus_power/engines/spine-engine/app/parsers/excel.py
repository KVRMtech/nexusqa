"""
Spine Engine — Excel Parser.

Parses Excel files (.xlsx/.xls) — typically rate tables and data extracts.
Uses openpyxl for real parsing, falls back to stub.
"""

from __future__ import annotations

import io

from ..models import ExtractedTable


class ExcelParser:
    """Parse Excel files (rate tables, data extracts)."""

    @staticmethod
    def parse(content: bytes, document_id: str) -> dict:
        try:
            import openpyxl
            return ExcelParser._parse_with_openpyxl(content, document_id)
        except ImportError:
            return ExcelParser._stub_parse(content, document_id)

    @staticmethod
    def _parse_with_openpyxl(content: bytes, document_id: str) -> dict:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = []
        tables = []
        full_text = ""

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                str_row = [str(c) if c is not None else "" for c in row]
                rows_data.append(str_row)

            if rows_data:
                headers = rows_data[0]
                data_rows = rows_data[1:]

                table = ExtractedTable(
                    document_id=document_id,
                    headers=headers,
                    rows=data_rows,
                    row_count=len(data_rows),
                    col_count=len(headers),
                    metadata={"sheet_name": sheet_name},
                )
                tables.append(table)

                # Convert to text
                text = f"Sheet: {sheet_name}\n"
                text += " | ".join(headers) + "\n"
                for row in data_rows[:100]:  # Limit for text representation
                    text += " | ".join(row) + "\n"
                full_text += text + "\n"

                sheets.append({
                    "sheet_name": sheet_name,
                    "row_count": len(data_rows),
                    "col_count": len(headers),
                    "headers": headers,
                })

        wb.close()

        return {
            "sheets": sheets,
            "tables": tables,
            "full_text": full_text,
            "page_count": len(sheets),
        }

    @staticmethod
    def _stub_parse(content: bytes, document_id: str) -> dict:
        return {
            "sheets": [{"sheet_name": "Sheet1", "row_count": 0, "col_count": 0, "headers": []}],
            "tables": [],
            "full_text": "[Stub Excel — install openpyxl for real parsing]",
            "page_count": 1,
        }
