"""
Excel workbook renderer — production-quality styled multi-sheet .xlsx.

Output format matches the exact specification:
  Sheet 1: "Test Cases"
    Columns: Test Case ID | Title | Step Number | Action | Expected Result
    - Merged rows for Test Case ID and Title (one row per step)
    - Professional header styling with dark blue background
    - Alternating row shading for readability
    - Data.X references highlighted in bold

  Sheet 2: "Preconditions"
    Columns: Test Case ID | # | Precondition

  Sheet 3: "Data Workbook"
    Columns: Test Case ID | FieldName | FieldValue | Type | Sensitive

  Sheet 4: "Summary"
    Metadata: export timestamp, total cases, coverage stats
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Sequence

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from nexus_sdk.models import ProductionTestCase


# ─── Style Constants ──────────────────────────────────────────────

# Header styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Alternating row colours
ROW_FILL_EVEN = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Data reference highlight  (Data.X)
DATA_REF_FONT = Font(name="Calibri", size=10, bold=True, color="0D47A1")

# Cell borders
THIN_BORDER = Border(
    left=Side(style="thin", color="B0BEC5"),
    right=Side(style="thin", color="B0BEC5"),
    top=Side(style="thin", color="B0BEC5"),
    bottom=Side(style="thin", color="B0BEC5"),
)

# Body cell defaults
BODY_FONT = Font(name="Calibri", size=10, color="212121")
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top")

# Summary sheet styling
SUMMARY_HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="37474F")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, color="212121")

# Sensitive data styling
SENSITIVE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
SENSITIVE_FONT = Font(name="Calibri", size=10, italic=True, color="E65100")


# ─── Public API ───────────────────────────────────────────────────

def render_excel(
    test_cases: Sequence[ProductionTestCase],
    output_path: Path,
    *,
    title: str = "Nexus QA — Test Cases",
    include_summary: bool = True,
) -> Path:
    """
    Render test cases to a production-quality Excel workbook.

    Args:
        test_cases: List of ProductionTestCase Pydantic models.
        output_path: Destination .xlsx file path.
        title: Workbook title (shown in Summary sheet).
        include_summary: Whether to add a Summary metadata sheet.

    Returns:
        The resolved output_path.

    Raises:
        ValueError: If test_cases is empty.
        OSError: If the output directory is not writable.
    """
    if not test_cases:
        raise ValueError("Cannot export empty test case list")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1 — Test Cases (steps)
    ws_cases = wb.active
    ws_cases.title = "Test Cases"  # type: ignore[union-attr]
    _build_test_cases_sheet(ws_cases, test_cases)  # type: ignore[arg-type]

    # Sheet 2 — Preconditions
    ws_pre = wb.create_sheet("Preconditions")
    _build_preconditions_sheet(ws_pre, test_cases)

    # Sheet 3 — Data Workbook
    ws_data = wb.create_sheet("Data Workbook")
    _build_data_workbook_sheet(ws_data, test_cases)

    # Sheet 4 — Summary
    if include_summary:
        ws_summary = wb.create_sheet("Summary")
        _build_summary_sheet(ws_summary, test_cases, title)

    wb.save(str(output_path))
    return output_path


# ─── Sheet Builders ───────────────────────────────────────────────

def _build_test_cases_sheet(
    ws: Worksheet,
    test_cases: Sequence[ProductionTestCase],
) -> None:
    """Build the main Test Cases sheet with Step-level rows."""
    headers = ["Test Case ID", "Title", "Step Number", "Action", "Expected Result"]
    _write_header_row(ws, headers)

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 50

    row_idx = 2  # start after header
    for tc_num, tc in enumerate(test_cases):
        if not tc.steps:
            # Test case with no steps — still show one row
            _write_body_cell(ws, row_idx, 1, tc.test_case_id, tc_num)
            _write_body_cell(ws, row_idx, 2, tc.title, tc_num)
            _write_body_cell(ws, row_idx, 3, "", tc_num)
            _write_body_cell(ws, row_idx, 4, "(No steps defined)", tc_num)
            _write_body_cell(ws, row_idx, 5, "", tc_num)
            row_idx += 1
            continue

        start_row = row_idx
        for step in tc.steps:
            _write_body_cell(ws, row_idx, 1, tc.test_case_id, tc_num)
            _write_body_cell(ws, row_idx, 2, tc.title, tc_num)
            _write_body_cell(ws, row_idx, 3, step.step_number, tc_num, center=True)
            _write_action_cell(ws, row_idx, 4, step.action, tc_num)
            _write_body_cell(ws, row_idx, 5, step.expected_result, tc_num)
            row_idx += 1

        # Merge ID and Title cells if multiple steps
        if len(tc.steps) > 1:
            end_row = row_idx - 1
            ws.merge_cells(
                start_row=start_row, start_column=1,
                end_row=end_row, end_column=1,
            )
            ws.merge_cells(
                start_row=start_row, start_column=2,
                end_row=end_row, end_column=2,
            )
            # Keep merged cell alignment
            ws.cell(start_row, 1).alignment = Alignment(
                vertical="center", horizontal="center", wrap_text=True,
            )
            ws.cell(start_row, 2).alignment = Alignment(
                vertical="center", wrap_text=True,
            )

    # Freeze top row + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{row_idx - 1}"


def _build_preconditions_sheet(
    ws: Worksheet,
    test_cases: Sequence[ProductionTestCase],
) -> None:
    """Build Preconditions sheet."""
    headers = ["Test Case ID", "#", "Precondition"]
    _write_header_row(ws, headers)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 80

    row_idx = 2
    for tc_num, tc in enumerate(test_cases):
        if not tc.preconditions:
            continue
        for i, pre in enumerate(tc.preconditions, 1):
            _write_body_cell(ws, row_idx, 1, tc.test_case_id, tc_num)
            _write_body_cell(ws, row_idx, 2, i, tc_num, center=True)
            _write_body_cell(ws, row_idx, 3, pre.description, tc_num)
            row_idx += 1

    ws.freeze_panes = "A2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:C{row_idx - 1}"


def _build_data_workbook_sheet(
    ws: Worksheet,
    test_cases: Sequence[ProductionTestCase],
) -> None:
    """Build the Data Workbook sheet (FieldName/FieldValue pairs)."""
    headers = ["Test Case ID", "FieldName", "FieldValue", "Type", "Sensitive"]
    _write_header_row(ws, headers)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    row_idx = 2
    for tc_num, tc in enumerate(test_cases):
        if not tc.data_workbook:
            continue
        for entry in tc.data_workbook:
            _write_body_cell(ws, row_idx, 1, tc.test_case_id, tc_num)
            _write_body_cell(ws, row_idx, 2, entry.field_name, tc_num)

            # Mask sensitive values
            if entry.is_sensitive:
                _write_sensitive_cell(ws, row_idx, 3, entry.field_value)
            else:
                _write_body_cell(ws, row_idx, 3, entry.field_value, tc_num)

            _write_body_cell(ws, row_idx, 4, entry.field_type, tc_num, center=True)
            _write_body_cell(
                ws, row_idx, 5, "Yes" if entry.is_sensitive else "No", tc_num, center=True,
            )
            row_idx += 1

    ws.freeze_panes = "A2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:E{row_idx - 1}"


def _build_summary_sheet(
    ws: Worksheet,
    test_cases: Sequence[ProductionTestCase],
    title: str,
) -> None:
    """Build a Summary/metadata sheet."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    now = datetime.now(timezone.utc)

    # Title
    ws.cell(1, 1, title).font = SUMMARY_HEADER_FONT
    ws.merge_cells("A1:B1")

    # Export metadata
    metadata_rows = [
        ("Export Date", now.strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Total Test Cases", len(test_cases)),
        ("Total Steps", sum(len(tc.steps) for tc in test_cases)),
        ("Total Data Fields", sum(len(tc.data_workbook) for tc in test_cases)),
        ("Total Preconditions", sum(len(tc.preconditions) for tc in test_cases)),
    ]

    # Count by type
    type_counts: dict[str, int] = {}
    priority_counts: dict[str, int] = {}
    status_counts: dict[str, int] = {}
    for tc in test_cases:
        type_counts[tc.test_type] = type_counts.get(tc.test_type, 0) + 1
        priority_counts[tc.priority] = priority_counts.get(tc.priority, 0) + 1
        status_counts[tc.status] = status_counts.get(tc.status, 0) + 1

    for label, value in metadata_rows:
        row = ws.max_row + 1 if ws.max_row > 1 else 3
        ws.cell(row, 1, label).font = SUMMARY_LABEL_FONT
        ws.cell(row, 2, value).font = SUMMARY_VALUE_FONT

    # Types breakdown
    row = ws.max_row + 2
    ws.cell(row, 1, "Breakdown by Type").font = SUMMARY_HEADER_FONT
    for t, c in sorted(type_counts.items()):
        row += 1
        ws.cell(row, 1, t.upper()).font = SUMMARY_LABEL_FONT
        ws.cell(row, 2, c).font = SUMMARY_VALUE_FONT

    # Priority breakdown
    row = ws.max_row + 2
    ws.cell(row, 1, "Breakdown by Priority").font = SUMMARY_HEADER_FONT
    for p, c in sorted(priority_counts.items()):
        row += 1
        ws.cell(row, 1, p.capitalize()).font = SUMMARY_LABEL_FONT
        ws.cell(row, 2, c).font = SUMMARY_VALUE_FONT

    # Status breakdown
    row = ws.max_row + 2
    ws.cell(row, 1, "Breakdown by Status").font = SUMMARY_HEADER_FONT
    for s, c in sorted(status_counts.items()):
        row += 1
        ws.cell(row, 1, s.capitalize()).font = SUMMARY_LABEL_FONT
        ws.cell(row, 2, c).font = SUMMARY_VALUE_FONT

    # Data coverage validation
    row = ws.max_row + 2
    ws.cell(row, 1, "Data Coverage Validation").font = SUMMARY_HEADER_FONT
    for tc in test_cases:
        errors = tc.validate_data_coverage()
        if errors:
            row += 1
            ws.cell(row, 1, tc.test_case_id).font = SUMMARY_LABEL_FONT
            ws.cell(row, 2, "; ".join(errors)).font = Font(
                name="Calibri", size=10, color="D32F2F",
            )


# ─── Cell Writers ─────────────────────────────────────────────────

def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write a styled header row at row 1."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _write_body_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: object,
    tc_index: int,
    *,
    center: bool = False,
) -> None:
    """Write a value cell with alternating row fill."""
    cell = ws.cell(row, col, value)
    cell.font = BODY_FONT
    cell.fill = ROW_FILL_EVEN if tc_index % 2 == 0 else ROW_FILL_ODD
    cell.alignment = CENTER_ALIGNMENT if center else BODY_ALIGNMENT
    cell.border = THIN_BORDER


_DATA_REF_PATTERN = re.compile(r"\(Data\.\w+\)")


def _write_action_cell(
    ws: Worksheet,
    row: int,
    col: int,
    action_text: str,
    tc_index: int,
) -> None:
    """
    Write an action cell; if it contains (Data.X) references,
    the entire cell gets the DATA_REF_FONT style (openpyxl
    doesn't support mixed-font in a single cell without
    rich-text, so we bold the whole cell when refs are present).
    """
    cell = ws.cell(row, col, action_text)
    cell.fill = ROW_FILL_EVEN if tc_index % 2 == 0 else ROW_FILL_ODD
    cell.alignment = BODY_ALIGNMENT
    cell.border = THIN_BORDER

    if _DATA_REF_PATTERN.search(action_text):
        cell.font = DATA_REF_FONT
    else:
        cell.font = BODY_FONT


def _write_sensitive_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: str,
) -> None:
    """Write a value marked as sensitive with special styling."""
    # Mask the actual value: show first 2 chars + asterisks
    if len(value) > 2:
        masked = value[:2] + "*" * min(len(value) - 2, 8)
    else:
        masked = "***"

    cell = ws.cell(row, col, masked)
    cell.font = SENSITIVE_FONT
    cell.fill = SENSITIVE_FILL
    cell.alignment = BODY_ALIGNMENT
    cell.border = THIN_BORDER
