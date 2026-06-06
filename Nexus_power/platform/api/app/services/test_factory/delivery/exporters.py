"""Excel / CSV / JSON exporters in the standard QA test-case format.

Column layout (exactly as specified):

    S.No | Test Case Name | Test Case Description | Test Steps | Test Data | Expected Result

One row per step; the Test Case Name + Description cells are merged across the
steps of a test case (Excel) so the sheet reads like a hand-authored test
plan.  Uses ``openpyxl`` directly (already a dependency) — full control over
the format, no dependency on a producer-specific contract.
"""

from __future__ import annotations

import csv
import io
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from nexus_sdk.models import ProductionTestCase

_BASE_HEADERS = [
    "S.No",
    "Test Case Name",
    "Test Case Description",
    "Test Steps",
    "Test Data",
    "Expected Result",
]
_OBSERVED_HEADER = "Observed in Recording"
_CONFIDENCE_HEADER = "Confidence"
_BASE_WIDTHS = [8, 34, 46, 62, 28, 50]
_OBSERVED_WIDTH = 42
_CONFIDENCE_WIDTH = 40

# Full layout (kept for callers/tests that want the column reference).
HEADERS = _BASE_HEADERS + [_OBSERVED_HEADER, _CONFIDENCE_HEADER]


def _headers(include_details: bool) -> list[str]:
    """Column layout — the technical columns are opt-in (role/toggle)."""
    extra = [_OBSERVED_HEADER, _CONFIDENCE_HEADER] if include_details else []
    return _BASE_HEADERS + extra


def _col_widths(include_details: bool) -> list[int]:
    extra = [_OBSERVED_WIDTH, _CONFIDENCE_WIDTH] if include_details else []
    return _BASE_WIDTHS + extra


def _confidence_cell(step) -> str:
    """Plain-language confidence for the export."""
    level = (getattr(step, "confidence", "") or "")
    reason = (getattr(step, "confidence_reason", "") or "")
    if not level:
        return ""
    if level == "high":
        return "Solid"
    return f"Review — {reason}" if reason else "Review"

# How the signal behind a step was sourced — keeps the evidence column honest.
_PROV_LABEL = {
    "demonstrated": "Observed",
    "available": "Option (captured)",
    "inferred": "Inferred",
}


def _observed_cell(step) -> str:
    """Compact evidence string captured in the recording, prefixed by provenance."""
    o = getattr(step, "observed", None) or {}
    prov = _PROV_LABEL.get(getattr(step, "provenance", "") or "", "")
    if o.get("url"):
        ev = f"navigate -> {o['url']}"
    elif o:
        tgt = f'"{o.get("label")}"' if o.get("label") else ""
        val = f' = "{o.get("value")}"' if o.get("value") else ""
        ev = f'{o.get("verb", "")} {tgt}{val}'.strip()
    else:
        ev = ""
    if prov and ev:
        return f"{prov}: {ev}"
    return prov or ev

EXPORT_MEDIA_TYPES = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv",
    "json": "application/json",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN = Side(style="thin", color="DDDDDD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _expected(step) -> str:
    return (getattr(step, "expected_result", None)
            or getattr(step, "expected", None) or "")


def _step_rows(tc: ProductionTestCase, include_details: bool = False) -> list[tuple]:
    """Return one row per step. The trailing 'Observed in Recording' cell is
    included only when ``include_details`` (role/toggle gated)."""
    rows: list[tuple] = []
    name = tc.name or ""
    description = tc.description or ""
    steps = tc.steps or []
    if not steps:
        base = (1, name, description, "(No steps defined)", "", "")
        return [base + (("", "") if include_details else ())]
    for idx, st in enumerate(steps, start=1):
        s_no = st.step_number if st.step_number is not None else idx
        base = (
            s_no,
            name,
            description,
            st.action or "",
            getattr(st, "data_ref", None) or "",
            _expected(st),
        )
        extra = (_observed_cell(st), _confidence_cell(st)) if include_details else ()
        rows.append(base + extra)
    return rows


def build_excel(test_cases: Sequence[ProductionTestCase], include_details: bool = False) -> bytes:
    """Render the suite to a styled .xlsx. The Observed column is opt-in."""
    headers = _headers(include_details)
    widths = _col_widths(include_details)
    ncols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = _BORDER
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    row = 2
    for tc in test_cases:
        start = row
        for record in _step_rows(tc, include_details):
            for col, value in enumerate(record, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1
        # Merge Name + Description across this test case's steps.
        if row - start > 1:
            for col in (2, 3):
                ws.merge_cells(start_row=start, start_column=col, end_row=row - 1, end_column=col)
                ws.cell(start, col).alignment = Alignment(vertical="top", wrap_text=True)

    for r in range(2, row):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.border = _BORDER
            if c not in (2, 3):
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(test_cases: Sequence[ProductionTestCase], include_details: bool = False) -> bytes:
    """Render the suite to CSV (UTF-8 BOM for Excel). Observed column is opt-in."""
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(_headers(include_details))
    for tc in test_cases:
        for row in _step_rows(tc, include_details):
            writer.writerow(list(row))
    return out.getvalue().encode("utf-8-sig")


def build_json(test_cases: Sequence[ProductionTestCase], include_details: bool = False) -> bytes:
    """Render the suite to JSON (full objects — the machine format always carries
    the observed evidence the script generator consumes)."""
    import json
    payload = {
        "version": "1.0",
        "test_cases": [tc.model_dump(mode="json") for tc in test_cases],
    }
    return json.dumps(payload, indent=2).encode("utf-8")
