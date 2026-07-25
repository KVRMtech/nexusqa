"""Phase R4 — reach formats (CSV / JUnit / XLSX / PDF), analytics, filters.

Spec: EXECUTION_EVIDENCE_REPORT_SPEC.md §2.13, §2.14, §2.16 (waves 2-3).

The honesty properties these pin:
  * rates are computed over EXECUTED cases only — a case that never ran can
    never inflate a pass rate;
  * the JUnit mapping keeps `<failure>` (the APPLICATION is at fault) separate
    from `<error>` (our automation is at fault), because collapsing them is
    exactly the blame-shifting the whole product refuses to do;
  * a FILTERED view never rewrites the totals, so it cannot read as a smaller,
    greener run;
  * PDF is generated without a third-party library so an air-gapped on-prem
    install can still produce an auditor's PDF.

Run from Nexus_power/platform/api:
    python -m pytest tests/test_evidence_report_r4.py -q
"""
from __future__ import annotations

import csv
import io
import os
import xml.etree.ElementTree as ET
import zipfile

from app.services.test_factory import report_formats as rf

_ROUTER = open(os.path.join(os.path.dirname(__file__), "..", "app", "routers",
                            "test_factory.py"), encoding="utf-8").read()


def _case(name, status, steps, *, executed=True, ttype="functional", dur=100):
    return {"test_case_id": f"tc-{name}", "name": name, "description": "",
            "test_type": ttype, "priority": "P1", "status": status,
            "executed": executed, "steps_declared": len(steps),
            "steps_executed": len(steps) if executed else 0,
            "counts": {}, "duration_ms": dur, "tags": [],
            "reproducibility": {}, "steps": steps,
            "not_executed_reason": "" if executed else "quarantined"}


def _step(n, status, *, ev="PROVEN", cause=None, cat=None, dur=10):
    return {"step_number": n, "status": status, "status_badge": "",
            "action": f"do thing {n}", "target": "getByRole('button')",
            "expected": "it works", "actual": "boom" if status != "passed" else "as expected",
            "duration_ms": dur, "evidence_class": ev, "oracle_provenance": {},
            "evidence": {"screenshot_url": "", "trace_url": ""},
            "analysis": ({"category": cat, "cause": cause, "detail": "d",
                          "evidence_quoted": [], "suggested": True} if cause else None)}


def _report() -> dict:
    good = _case("Verify user can complete the 'apply' flow", "passed",
                 [_step(1, "passed"), _step(2, "passed", dur=50)])
    defect = _case("Verify validation on 'apply'", "completed_with_defects",
                   [_step(1, "passed"),
                    _step(2, "defect_found", ev="UNVERIFIED",
                          cause="validation_missing", cat="application_defect", dur=900)])
    broken = _case("Verify user can navigate to 'claims'", "execution_error",
                   [_step(1, "execution_error", ev="UNVERIFIED",
                          cause="ambiguous_locator", cat="product_script_defect")],
                   ttype="combination")
    never = _case("Verify combination rank 9", "not_executed", [], executed=False,
                  ttype="combination", dur=0)
    return {"report_version": "1.0", "generated_at": "2026-07-25T00:00:00+00:00",
            "run": {"run_id": "r1", "environment": "certification",
                    "started_at": "2026-07-25T00:00:00+00:00",
                    "ingested_totals": {"total_steps": 5}},
            "trust": {"certified": True, "certification_run": {"run_id": "c1"},
                      "quarantined_count": 1, "uncertified_exploratory_count": 0,
                      "statement": "certified first", "quarantined": [],
                      "uncertified_exploratory": [], "suite_size": 4,
                      "oracle_scorecard": None},
            "summary": {"artifact_id": "a1", "total_flows": 2,
                        "total_cases_generated": 4, "total_cases_executed": 3,
                        "total_steps_executed": 5,
                        "case_counts": {"passed": 1, "defect_found": 1,
                                        "execution_error": 1, "blocked": 0,
                                        "needs_review": 0, "skipped": 0,
                                        "cancelled": 0, "not_executed": 1, "total": 4},
                        "step_counts": {"passed": 3, "defect_found": 1,
                                        "execution_error": 1, "blocked": 0,
                                        "needs_review": 0, "skipped": 0,
                                        "cancelled": 0, "total": 5}},
            "flows": [
                {"flow_key": "apply", "flow_label": "apply",
                 "cases": [good, defect, never], "case_count": 3,
                 "duration_ms": 200, "pass_percentage": 50.0, "defect_count": 1,
                 "counts": {}},
                {"flow_key": "claims", "flow_label": "claims", "cases": [broken],
                 "case_count": 1, "duration_ms": 100, "pass_percentage": 0.0,
                 "defect_count": 0, "counts": {}}],
            "defects": {"unique_defects": 2, "total_occurrences": 5,
                        "by_lifecycle": {"open": 2}, "window_runs": 3,
                        "defects": [
                            {"signature": "sig1", "scenario_id": "tc1", "case_name": "c1",
                             "step_number": 2, "display_status": "defect_found",
                             "category": "application_defect", "cause": "validation_missing",
                             "lifecycle": "open", "occurrence_count": 3,
                             "first_seen": "a", "last_seen": "b", "fingerprint": "fp1",
                             "occurrences": []},
                            {"signature": "sig2", "scenario_id": "tc2", "case_name": "c2",
                             "step_number": 1, "display_status": "execution_error",
                             "category": "product_script_defect", "cause": "ambiguous_locator",
                             "lifecycle": "open", "occurrence_count": 2,
                             "first_seen": "a", "last_seen": "b", "fingerprint": "fp2",
                             "occurrences": []}],
                        "note": "n"},
            "diff": {"available": False, "reason": "none"},
            "coverage": {"note": "un-run cases are listed, never counted green",
                         "cases_not_executed": [{"test_case_id": "tc-x", "name": "n",
                                                 "reason": "quarantined"}],
                         "cases_not_executed_count": 1, "quarantined_count": 1,
                         "uncertified_exploratory_count": 0},
            "doctrine": {}}


# ── §2.14 analytics ──────────────────────────────────────────────────────────

def test_rates_are_over_executed_cases_only():
    a = rf.build_analytics(_report())
    assert a["cases_executed"] == 3          # the not_executed case is excluded
    assert a["pass_rate_pct"] == 33.3        # 1 of 3, NOT 1 of 4
    assert a["defect_rate_pct"] == 33.3
    assert a["execution_error_rate_pct"] == 33.3


def test_unrun_cases_can_never_inflate_the_pass_rate():
    rep = _report()
    a1 = rf.build_analytics(rep)
    # add 10 more never-executed cases; the pass rate must not move
    rep["flows"][0]["cases"] += [
        _case(f"never {i}", "not_executed", [], executed=False) for i in range(10)]
    assert rf.build_analytics(rep)["pass_rate_pct"] == a1["pass_rate_pct"]


def test_analytics_reports_evidence_class_not_a_confidence_score():
    a = rf.build_analytics(_report())
    dist = a["evidence_class_distribution"]
    assert set(dist) <= {"PROVEN", "INFERRED", "UNVERIFIED"}
    assert "confidence" not in " ".join(a.keys()).lower()


def test_slowest_step_and_longest_case_are_identified():
    a = rf.build_analytics(_report())
    assert a["slowest_step"]["duration_ms"] == 900
    assert a["longest_case"]["duration_ms"] == 100


def test_most_frequent_findings_keep_their_class():
    a = rf.build_analytics(_report())
    keys = [x["key"] for x in a["most_frequent_findings"]]
    assert any(k.startswith("defect_found:") for k in keys)
    assert any(k.startswith("execution_error:") for k in keys)


# ── §2.13 filters ────────────────────────────────────────────────────────────

def test_filter_by_status():
    out = rf.filter_report(_report(), statuses={"completed_with_defects"})
    names = [c["name"] for f in out["flows"] for c in f["cases"]]
    assert names == ["Verify validation on 'apply'"]


def test_filter_by_flow_and_type_and_search():
    assert rf.filter_report(_report(), flow="claims")["filter"]["matched_cases"] == 1
    assert rf.filter_report(_report(), test_type="combination")["filter"]["matched_cases"] == 2
    assert rf.filter_report(_report(), search="validation")["filter"]["matched_cases"] == 1


def test_filtered_view_does_not_rewrite_the_totals():
    """A filtered report must not be mistakable for a smaller, greener run."""
    out = rf.filter_report(_report(), statuses={"passed"})
    assert out["summary"]["total_cases_generated"] == 4
    assert out["summary"]["case_counts"]["execution_error"] == 1
    assert "still describe the whole execution" in out["filter"]["note"]


# ── CSV ──────────────────────────────────────────────────────────────────────

def test_csv_has_one_row_per_step_plus_unrun_cases():
    rows = list(csv.reader(io.StringIO(rf.to_csv(_report()))))
    header, body = rows[0], rows[1:]
    assert header[0] == "flow" and "evidence_class" in header
    assert len(body) == 6          # 5 executed steps + 1 not-executed case row
    assert any(r[7] == "not_executed" for r in body)


def test_csv_carries_attribution_and_ai_suggested_flag():
    rows = list(csv.DictReader(io.StringIO(rf.to_csv(_report()))))
    bad = [r for r in rows if r["step_status"] == "defect_found"][0]
    assert bad["attribution_category"] == "application_defect"
    assert bad["ai_suggested"] == "True"


# ── JUnit XML ────────────────────────────────────────────────────────────────

def test_junit_separates_application_failure_from_our_execution_error():
    """THE distinction plain JUnit consumers otherwise lose."""
    root = ET.fromstring(rf.to_junit_xml(_report()))
    suite = root.find("testsuite")
    by_name = {tc.get("name"): tc for tc in suite.findall("testcase")}
    defect = by_name["Verify validation on 'apply'"]
    broken = by_name["Verify user can navigate to 'claims'"]
    assert defect.find("failure") is not None and defect.find("error") is None
    assert broken.find("error") is not None and broken.find("failure") is None
    assert defect.find("failure").get("type") == "application_defect"


def test_junit_counts_and_carries_the_precise_class_in_properties():
    xml = rf.to_junit_xml(_report())
    root = ET.fromstring(xml)
    assert root.get("failures") == "1" and root.get("errors") == "1"
    assert root.get("skipped") == "1" and root.get("tests") == "4"
    tc = root.find("testsuite").findall("testcase")[0]
    props = {p.get("name"): p.get("value") for p in tc.find("properties")}
    assert props["nexus.status"] == "passed"
    assert "Lossy by design" in xml           # the mapping declares itself


def test_junit_skipped_carries_the_reason():
    root = ET.fromstring(rf.to_junit_xml(_report()))
    sk = [tc for tc in root.find("testsuite").findall("testcase")
          if tc.find("skipped") is not None][0]
    assert "quarantined" in sk.find("skipped").get("message")


# ── XLSX ─────────────────────────────────────────────────────────────────────

def test_xlsx_is_a_real_workbook_with_the_expected_sheets():
    blob = rf.to_xlsx(_report())
    assert blob[:2] == b"PK"
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        assert "xl/workbook.xml" in z.namelist()
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob))
    assert wb.sheetnames == ["Summary", "Cases", "Steps", "Defects"]
    assert wb["Steps"].max_row == 6           # header + 5 steps
    assert wb["Defects"].max_row == 3         # header + 2 defects


def test_xlsx_summary_lists_every_bucket_including_zeros():
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(rf.to_xlsx(_report())))
    labels = [r[0] for r in wb["Summary"].iter_rows(values_only=True)]
    for bucket in ("passed", "defect_found", "execution_error", "blocked",
                   "needs_review", "skipped", "cancelled"):
        assert bucket in labels


# ── PDF ──────────────────────────────────────────────────────────────────────

def test_pdf_is_a_valid_document_without_third_party_libraries():
    blob = rf.to_pdf(_report())
    assert blob.startswith(b"%PDF-1.4")
    assert blob.rstrip().endswith(b"%%EOF")
    assert b"/Type /Catalog" in blob and b"/Type /Page" in blob
    assert b"xref" in blob and b"trailer" in blob
    assert len(blob) > 900


def test_pdf_module_imports_no_pdf_dependency():
    """Assert on IMPORTS, not bare strings — the rationale comment legitimately
    names the libraries we deliberately avoid."""
    src = open(rf.__file__, encoding="utf-8").read()
    for lib in ("reportlab", "weasyprint", "fpdf", "xhtml2pdf"):
        assert f"import {lib}" not in src, "PDF must work on an air-gapped install"
        assert f"from {lib}" not in src


def test_pdf_lists_every_case_and_paginates():
    """An audit PDF that never says WHICH cases ran is not an audit artifact —
    so the page count must grow with the suite."""
    def page_count(blob: bytes) -> int:
        i = blob.index(b"/Type /Pages")
        return int(blob[i:i + 160].split(b"/Count ")[1].split(b" ")[0].strip(b" >>\n"))

    small = rf.to_pdf(_report())
    rep = _report()
    rep["flows"][0]["cases"] += [
        _case(f"case {i}", "passed", [_step(1, "passed")]) for i in range(120)]
    big = rf.to_pdf(rep)
    # Measure PAGES, not bytes: the content stream is zlib-compressed and 120
    # near-identical case lines compress away to almost nothing, so a byte-size
    # threshold would test the compressor rather than the pagination.
    assert page_count(small) >= 1
    assert page_count(big) >= 5, "124 cases must span several pages"
    assert page_count(big) > page_count(small)
    assert len(big) > len(small)


def test_pdf_names_cases_that_did_not_execute():
    import zlib
    blob = rf.to_pdf(_report())
    text = b""
    for chunk in blob.split(b"stream\n")[1:]:
        raw = chunk.split(b"\nendstream")[0]
        try:
            text += zlib.decompress(raw)
        except Exception:
            pass
    assert b"NOT EXECUTED" in text
    assert b"Test cases" in text


# ── routes ───────────────────────────────────────────────────────────────────

def test_format_route_exists_is_rbac_gated_redacted_and_audited():
    assert '/report/export.{fmt}' in _ROUTER
    assert "_export_role_ok(user)" in _ROUTER
    assert "report_export.redact_report(report)" in _ROUTER
    assert 'event_type="evidence_exported"' in _ROUTER


def test_unknown_format_is_rejected():
    assert "unknown format" in _ROUTER


def test_analytics_route_exists():
    assert "/report/analytics" in _ROUTER
